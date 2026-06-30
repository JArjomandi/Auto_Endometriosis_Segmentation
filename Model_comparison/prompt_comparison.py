from pathlib import Path
import ast
import re

import numpy as np
import pandas as pd
from PIL import Image

import matplotlib.pyplot as plt

from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


# =============================================================================
# Paths
# =============================================================================

RESULTS_ROOT = Path(r"F:\Results\SAM_Benchmarking")

OUTPUT_ROOT = RESULTS_ROOT / "Model_comparison" / "SAM_prompt_comparison"
FIGURE_DIR = OUTPUT_ROOT / "figures_600dpi"
SUMMARY_XLSX = OUTPUT_ROOT / "sam_prompt_comparison_summary.xlsx"

DISCOVERED_CSV = OUTPUT_ROOT / "discovered_prompt_files.csv"
VALUES_CSV = OUTPUT_ROOT / "calculated_prompt_metrics_image_level.csv"
CHECK_CSV = OUTPUT_ROOT / "metric_calculation_check.csv"


# =============================================================================
# Dataset paths
# =============================================================================

STANDARDIZED_DATASETS = {
    "ENID": Path(r"F:\Datasets\Standardized datasets\ENID\ENID 60_20_20 Split"),
    "GLENDA": Path(r"F:\Datasets\Standardized datasets\GLENDA\GLENDA 60_20_20 split"),
    "GLENDA_clean": Path(r"F:\Datasets\Standardized datasets\GLENDA_clean\GLENDA_clean 60_20_20 split"),
}


# =============================================================================
# Datasets / splits / models
# =============================================================================

DATASETS = [
    "ENID",
    "GLENDA",
    "GLENDA_clean",
]

SPLITS = [
    "val",
    "test",
]

SAM_MODEL_FOLDERS = {
    "SAM2": [
        "SAM2",
    ],
    "MedSAM": [
        "MedSAM",
    ],
    "SAM-Med2D": [
        "SAM-Med2D",
        "SAM_Med2D",
        "SAMMed2D",
    ],
    "SurgiSAM": [
        "SurgiSAM2",
        "SurgiSAM",
    ],
}

SAM_MODELS = list(SAM_MODEL_FOLDERS.keys())


# =============================================================================
# Prompt folders
# =============================================================================

PROMPT_TYPES = [
    "Point",
    "Box",
    "Box+Point",
    "Box+Point+Negative",
]

PROMPT_FOLDER_MAP = {
    "Point": [
        "GT_point",
    ],
    "Box": [
        "GT_box",
    ],
    "Box+Point": [
        "GT_box_point",
        "GT_point_box",
    ],
    "Box+Point+Negative": [
        "GT_box_posneg",
        "GT_box_pos_neg",
        "GT_box_point_negative",
        "GT_box_point_neg",
    ],
}

PROMPT_SUPPORT = {
    "SAM2": {
        "Point": True,
        "Box": True,
        "Box+Point": True,
        "Box+Point+Negative": True,
    },
    "MedSAM": {
        "Point": False,
        "Box": True,
        "Box+Point": False,
        "Box+Point+Negative": False,
    },
    "SAM-Med2D": {
        "Point": True,
        "Box": True,
        "Box+Point": False,
        "Box+Point+Negative": False,
    },
    "SurgiSAM": {
        "Point": True,
        "Box": True,
        "Box+Point": True,
        "Box+Point+Negative": True,
    },
}


# =============================================================================
# Metrics
# =============================================================================

METRICS = {
    "dice": {
        "title": "Dice Scores",
        "ylabel": "Dice score",
    },
    "iou": {
        "title": "IoU Scores",
        "ylabel": "IoU score",
    },
    "precision": {
        "title": "Precision Scores",
        "ylabel": "Precision",
    },
    "recall": {
        "title": "Recall Scores",
        "ylabel": "Recall",
    },
}


# =============================================================================
# Plot settings
# =============================================================================

DPI = 600

# Compact figure
FIG_WIDTH = 9.8
FIG_HEIGHT = 5.4

TITLE_FONT_SIZE = 14
AXIS_FONT_SIZE = 13

# Main prompt tick labels: smaller but bold
TICK_FONT_SIZE = 13

# Skewed model labels under each box
MODEL_LABEL_FONT_SIZE = 12

# Compact prompt groups, with slight white space between boxes
BOX_WIDTH = 0.095
GROUP_GAP = 0.68

BOX_COLOR = "#00BDD6"
MEAN_COLOR = "red"
MEDIAN_COLOR = "black"


# =============================================================================
# Basic helpers
# =============================================================================

def normalize_name(value):
    value = str(value).strip().lower()
    value = value.replace("+", "_")
    value = re.sub(r"[^a-z0-9]+", "_", value)
    value = re.sub(r"_+", "_", value)
    return value.strip("_")


def find_existing_model_folder(dataset_key, model_display_name):
    dataset_root = RESULTS_ROOT / dataset_key

    for folder_name in SAM_MODEL_FOLDERS[model_display_name]:
        candidate = dataset_root / folder_name
        if candidate.exists() and candidate.is_dir():
            return candidate

    return None


def find_file_by_name(folder: Path, file_name: str):
    direct = folder / file_name
    if direct.exists():
        return direct

    stem = Path(file_name).stem
    extensions = [
        Path(file_name).suffix,
        ".png",
        ".jpg",
        ".jpeg",
        ".tif",
        ".tiff",
        ".bmp",
    ]

    for ext in extensions:
        if ext == "":
            continue

        candidate = folder / f"{stem}{ext}"
        if candidate.exists():
            return candidate

    matches = list(folder.rglob(file_name))
    if len(matches) > 0:
        return matches[0]

    for ext in extensions:
        if ext == "":
            continue

        matches = list(folder.rglob(f"{stem}{ext}"))
        if len(matches) > 0:
            return matches[0]

    return None


def read_binary_mask(mask_path: Path):
    mask = Image.open(mask_path).convert("L")
    mask_np = np.array(mask)
    return (mask_np > 0).astype(np.uint8)


def resize_mask_to_shape(mask_np: np.ndarray, target_shape):
    target_h, target_w = target_shape[:2]

    if mask_np.shape[:2] == (target_h, target_w):
        return (mask_np > 0).astype(np.uint8)

    mask_img = Image.fromarray((mask_np > 0).astype(np.uint8) * 255)
    mask_img = mask_img.resize((target_w, target_h), resample=Image.Resampling.NEAREST)

    return (np.array(mask_img) > 0).astype(np.uint8)


def parse_bbox(value):
    if value is None:
        return None

    if isinstance(value, float) and pd.isna(value):
        return None

    if isinstance(value, (list, tuple)) and len(value) == 4:
        return [float(v) for v in value]

    value = str(value).strip()

    if value == "" or value.lower() == "nan":
        return None

    try:
        parsed = ast.literal_eval(value)
        if isinstance(parsed, (list, tuple)) and len(parsed) == 4:
            return [float(v) for v in parsed]
    except Exception:
        pass

    return None


# =============================================================================
# Connected components for GT lesion extraction
# =============================================================================

def connected_components(binary_mask: np.ndarray):
    binary_mask = (binary_mask > 0).astype(np.uint8)

    try:
        from scipy import ndimage

        labeled, num_components = ndimage.label(binary_mask > 0)

        components = []
        for component_id in range(1, num_components + 1):
            comp = (labeled == component_id).astype(np.uint8)
            if comp.sum() > 0:
                components.append(comp)

        return components

    except Exception:
        pass

    if binary_mask.sum() > 0:
        return [binary_mask]

    return []


def get_gt_component_from_bbox_or_lesion_id(gt_mask: np.ndarray, row):
    gt_mask = (gt_mask > 0).astype(np.uint8)

    components = connected_components(gt_mask)

    if len(components) == 0:
        return gt_mask

    bbox = None

    if "bbox_xyxy" in row.index:
        bbox = parse_bbox(row["bbox_xyxy"])

    if bbox is None:
        needed = ["bbox_x1", "bbox_y1", "bbox_x2", "bbox_y2"]
        if all(column in row.index for column in needed):
            try:
                bbox = [
                    float(row["bbox_x1"]),
                    float(row["bbox_y1"]),
                    float(row["bbox_x2"]),
                    float(row["bbox_y2"]),
                ]
            except Exception:
                bbox = None

    if bbox is not None:
        h, w = gt_mask.shape[:2]

        x1, y1, x2, y2 = bbox

        x1 = max(0, min(w - 1, int(round(x1))))
        x2 = max(0, min(w - 1, int(round(x2))))
        y1 = max(0, min(h - 1, int(round(y1))))
        y2 = max(0, min(h - 1, int(round(y2))))

        if x2 < x1:
            x1, x2 = x2, x1

        if y2 < y1:
            y1, y2 = y2, y1

        bbox_mask = np.zeros_like(gt_mask, dtype=np.uint8)
        bbox_mask[y1:y2 + 1, x1:x2 + 1] = 1

        best_component = None
        best_overlap = -1

        for comp in components:
            overlap = int(np.logical_and(comp > 0, bbox_mask > 0).sum())
            if overlap > best_overlap:
                best_overlap = overlap
                best_component = comp

        if best_component is not None and best_overlap > 0:
            return best_component.astype(np.uint8)

    if "lesion_id" in row.index:
        try:
            lesion_id = int(row["lesion_id"])
            if 0 <= lesion_id < len(components):
                return components[lesion_id].astype(np.uint8)
        except Exception:
            pass

    return gt_mask


# =============================================================================
# Metric calculation
# =============================================================================

def calculate_metrics(pred_mask: np.ndarray, gt_mask: np.ndarray):
    pred_mask = (pred_mask > 0).astype(np.uint8)
    gt_mask = (gt_mask > 0).astype(np.uint8)

    if pred_mask.shape != gt_mask.shape:
        pred_mask = resize_mask_to_shape(pred_mask, gt_mask.shape)

    pred = pred_mask > 0
    gt = gt_mask > 0

    tp = int(np.logical_and(pred, gt).sum())
    fp = int(np.logical_and(pred, ~gt).sum())
    fn = int(np.logical_and(~pred, gt).sum())

    pred_sum = int(pred.sum())
    gt_sum = int(gt.sum())

    dice_den = 2 * tp + fp + fn
    iou_den = tp + fp + fn
    precision_den = tp + fp
    recall_den = tp + fn

    dice = 1.0 if dice_den == 0 else (2.0 * tp) / dice_den
    iou = 1.0 if iou_den == 0 else tp / iou_den
    precision = 1.0 if precision_den == 0 else tp / precision_den
    recall = 1.0 if recall_den == 0 else tp / recall_den

    return {
        "dice": float(dice),
        "iou": float(iou),
        "precision": float(precision),
        "recall": float(recall),
        "tp": tp,
        "fp": fp,
        "fn": fn,
        "pred_area": pred_sum,
        "gt_area": gt_sum,
    }


# =============================================================================
# File discovery
# =============================================================================

def discover_prompt_files():
    rows = []

    for dataset_key in DATASETS:
        for model_name in SAM_MODELS:
            model_root = find_existing_model_folder(dataset_key, model_name)

            if model_root is None:
                rows.append(
                    {
                        "Dataset": dataset_key,
                        "Split": "",
                        "Model": model_name,
                        "Prompt type": "",
                        "Prompt folder": "",
                        "Supported": "",
                        "Status": "Model folder missing",
                        "CSV path": "",
                    }
                )
                continue

            frozen_root = model_root / "frozen"

            for prompt_type in PROMPT_TYPES:
                supported = PROMPT_SUPPORT[model_name][prompt_type]

                if not supported:
                    for split_key in SPLITS:
                        rows.append(
                            {
                                "Dataset": dataset_key,
                                "Split": split_key,
                                "Model": model_name,
                                "Prompt type": prompt_type,
                                "Prompt folder": "",
                                "Supported": "No / not standard",
                                "Status": "Skipped - prompt not supported/standard",
                                "CSV path": "",
                            }
                        )
                    continue

                found_any_folder = False

                for prompt_folder_name in PROMPT_FOLDER_MAP[prompt_type]:
                    prompt_folder = frozen_root / prompt_folder_name

                    if not prompt_folder.exists():
                        continue

                    found_any_folder = True

                    for split_key in SPLITS:
                        csv_path = prompt_folder / split_key / "inference_results.csv"

                        rows.append(
                            {
                                "Dataset": dataset_key,
                                "Split": split_key,
                                "Model": model_name,
                                "Prompt type": prompt_type,
                                "Prompt folder": prompt_folder_name,
                                "Supported": "Yes",
                                "Status": "OK" if csv_path.exists() else "Missing inference_results.csv",
                                "CSV path": str(csv_path),
                            }
                        )

                if not found_any_folder:
                    for split_key in SPLITS:
                        rows.append(
                            {
                                "Dataset": dataset_key,
                                "Split": split_key,
                                "Model": model_name,
                                "Prompt type": prompt_type,
                                "Prompt folder": "",
                                "Supported": "Yes",
                                "Status": "Prompt folder missing",
                                "CSV path": "",
                            }
                        )

    return pd.DataFrame(rows)


# =============================================================================
# Metric collection from masks
# =============================================================================

def get_gt_mask_name(row):
    for col in ["mask_name", "gt_mask_name", "gt_name", "label_name"]:
        if col in row.index:
            value = str(row[col])
            if value and value.lower() != "nan":
                return value
    return None


def get_pred_mask_name(row):
    for col in [
        "instance_mask_name",
        "prediction_name",
        "pred_mask_name",
        "mask_prediction_name",
    ]:
        if col in row.index:
            value = str(row[col])
            if value and value.lower() != "nan":
                return value
    return None


def find_prediction_mask(csv_path: Path, pred_mask_name: str):
    split_dir = csv_path.parent

    candidate_dirs = [
        split_dir,
        split_dir / "instance_masks",
        split_dir / "predicted_instance_masks",
        split_dir / "prediction_masks",
        split_dir / "predictions",
        split_dir / "masks",
        split_dir / "binary_masks",
        split_dir / "output_masks",
    ]

    for folder in candidate_dirs:
        if folder.exists():
            found = find_file_by_name(folder, pred_mask_name)
            if found is not None:
                return found

    found = find_file_by_name(split_dir, pred_mask_name)
    return found


def find_gt_mask(dataset_key: str, split_key: str, gt_mask_name: str):
    dataset_root = STANDARDIZED_DATASETS[dataset_key]
    mask_folder = dataset_root / split_key / "masks"

    found = find_file_by_name(mask_folder, gt_mask_name)
    return found


def collect_metrics_from_csv(file_row):
    dataset_key = file_row["Dataset"]
    split_key = file_row["Split"]
    model_name = file_row["Model"]
    prompt_type = file_row["Prompt type"]
    prompt_folder = file_row["Prompt folder"]
    csv_path = Path(file_row["CSV path"])

    metric_rows = []
    check_rows = []

    try:
        df = pd.read_csv(csv_path)
    except Exception as error:
        check_rows.append(
            {
                "Dataset": dataset_key,
                "Split": split_key,
                "Model": model_name,
                "Prompt type": prompt_type,
                "Prompt folder": prompt_folder,
                "Status": f"Could not read inference_results.csv: {error}",
                "CSV path": str(csv_path),
            }
        )
        return metric_rows, check_rows

    for row_index, row in df.iterrows():
        gt_mask_name = get_gt_mask_name(row)
        pred_mask_name = get_pred_mask_name(row)

        if gt_mask_name is None:
            check_rows.append(
                {
                    "Dataset": dataset_key,
                    "Split": split_key,
                    "Model": model_name,
                    "Prompt type": prompt_type,
                    "Prompt folder": prompt_folder,
                    "Row index": row_index,
                    "Status": "Missing GT mask name column/value",
                    "CSV path": str(csv_path),
                }
            )
            continue

        if pred_mask_name is None:
            check_rows.append(
                {
                    "Dataset": dataset_key,
                    "Split": split_key,
                    "Model": model_name,
                    "Prompt type": prompt_type,
                    "Prompt folder": prompt_folder,
                    "Row index": row_index,
                    "Status": "Missing prediction mask name column/value",
                    "CSV path": str(csv_path),
                }
            )
            continue

        gt_path = find_gt_mask(dataset_key, split_key, gt_mask_name)
        pred_path = find_prediction_mask(csv_path, pred_mask_name)

        if gt_path is None:
            check_rows.append(
                {
                    "Dataset": dataset_key,
                    "Split": split_key,
                    "Model": model_name,
                    "Prompt type": prompt_type,
                    "Prompt folder": prompt_folder,
                    "Row index": row_index,
                    "Status": f"GT mask not found: {gt_mask_name}",
                    "CSV path": str(csv_path),
                }
            )
            continue

        if pred_path is None:
            check_rows.append(
                {
                    "Dataset": dataset_key,
                    "Split": split_key,
                    "Model": model_name,
                    "Prompt type": prompt_type,
                    "Prompt folder": prompt_folder,
                    "Row index": row_index,
                    "Status": f"Prediction mask not found: {pred_mask_name}",
                    "CSV path": str(csv_path),
                }
            )
            continue

        try:
            gt_mask_full = read_binary_mask(gt_path)
            pred_mask = read_binary_mask(pred_path)
            gt_component = get_gt_component_from_bbox_or_lesion_id(gt_mask_full, row)

            metrics = calculate_metrics(pred_mask, gt_component)

        except Exception as error:
            check_rows.append(
                {
                    "Dataset": dataset_key,
                    "Split": split_key,
                    "Model": model_name,
                    "Prompt type": prompt_type,
                    "Prompt folder": prompt_folder,
                    "Row index": row_index,
                    "Status": f"Metric calculation failed: {error}",
                    "CSV path": str(csv_path),
                }
            )
            continue

        image_name = str(row["image_name"]) if "image_name" in row.index else ""
        lesion_id = row["lesion_id"] if "lesion_id" in row.index else ""

        base_info = {
            "Dataset": dataset_key,
            "Split": split_key,
            "Model": model_name,
            "Prompt type": prompt_type,
            "Prompt folder": prompt_folder,
            "image_name": image_name,
            "gt_mask_name": gt_mask_name,
            "pred_mask_name": pred_mask_name,
            "lesion_id": lesion_id,
            "GT path": str(gt_path),
            "Prediction path": str(pred_path),
            "CSV path": str(csv_path),
        }

        for metric_key in METRICS.keys():
            metric_rows.append(
                {
                    **base_info,
                    "Metric": metric_key,
                    "Value": metrics[metric_key],
                    "tp": metrics["tp"],
                    "fp": metrics["fp"],
                    "fn": metrics["fn"],
                    "pred_area": metrics["pred_area"],
                    "gt_area": metrics["gt_area"],
                }
            )

        check_rows.append(
            {
                "Dataset": dataset_key,
                "Split": split_key,
                "Model": model_name,
                "Prompt type": prompt_type,
                "Prompt folder": prompt_folder,
                "Row index": row_index,
                "Status": "OK",
                "CSV path": str(csv_path),
            }
        )

    return metric_rows, check_rows


def collect_all_metrics(discovered_df):
    all_metric_rows = []
    all_check_rows = []

    ok_files = discovered_df[discovered_df["Status"] == "OK"].copy()

    print(f"OK inference_results.csv files found: {len(ok_files)}")

    for _, file_row in ok_files.iterrows():
        print(
            f"Calculating: {file_row['Dataset']} | {file_row['Split']} | "
            f"{file_row['Model']} | {file_row['Prompt type']} | {file_row['CSV path']}"
        )

        metric_rows, check_rows = collect_metrics_from_csv(file_row)

        all_metric_rows.extend(metric_rows)
        all_check_rows.extend(check_rows)

    metrics_df = pd.DataFrame(all_metric_rows)
    check_df = pd.DataFrame(all_check_rows)

    return metrics_df, check_df


# =============================================================================
# Summary tables
# =============================================================================

def make_prompt_support_df():
    rows = []

    for model_name in SAM_MODELS:
        row = {"Model used": model_name}

        for prompt_type in PROMPT_TYPES:
            row[prompt_type] = (
                "Yes"
                if PROMPT_SUPPORT[model_name][prompt_type]
                else "No / not standard"
            )

        rows.append(row)

    return pd.DataFrame(rows)


def summarize_values(values_df):
    if len(values_df) == 0:
        return pd.DataFrame()

    summary_df = (
        values_df
        .groupby(
            ["Dataset", "Split", "Model", "Prompt type", "Prompt folder", "Metric"],
            as_index=False,
        )
        .agg(
            N=("Value", "count"),
            Mean=("Value", "mean"),
            Std=("Value", "std"),
            Median=("Value", "median"),
            Q1=("Value", lambda x: x.quantile(0.25)),
            Q3=("Value", lambda x: x.quantile(0.75)),
            Min=("Value", "min"),
            Max=("Value", "max"),
        )
    )

    summary_df["Std"] = summary_df["Std"].fillna(0.0)

    summary_df["Mean ± SD"] = summary_df.apply(
        lambda row: f"{row['Mean']:.4f} ± {row['Std']:.4f}",
        axis=1,
    )

    ordered = [
        "Dataset",
        "Split",
        "Model",
        "Prompt type",
        "Prompt folder",
        "Metric",
        "N",
        "Mean ± SD",
        "Mean",
        "Std",
        "Median",
        "Q1",
        "Q3",
        "Min",
        "Max",
    ]

    return summary_df[ordered]


def make_wide_summary(summary_df):
    if len(summary_df) == 0:
        return pd.DataFrame()

    wide_df = summary_df.pivot_table(
        index=["Dataset", "Split", "Metric", "Prompt type"],
        columns="Model",
        values="Mean ± SD",
        aggfunc="first",
    ).reset_index()

    return wide_df


# =============================================================================
# Plotting
# =============================================================================

def get_offsets_for_models(n_models):
    """
    Controls spacing between boxes above each prompt tick.

    Requirements:
    - 2-box groups closer together.
    - 3/4-box groups still have small visible white space.
    """
    if n_models == 1:
        return [0.0]

    if n_models == 2:
        return [-0.070, 0.070]

    if n_models == 3:
        return [-0.145, 0.0, 0.145]

    if n_models == 4:
        return [-0.210, -0.070, 0.070, 0.210]

    return np.linspace(-0.210, 0.210, n_models)


def make_single_boxplot(values_df, dataset_key, split_key, metric_key):
    plot_df = values_df[
        (values_df["Dataset"] == dataset_key)
        & (values_df["Split"] == split_key)
        & (values_df["Metric"] == metric_key)
    ].copy()

    if len(plot_df) == 0:
        print(f"No data for plot: {dataset_key} | {split_key} | {metric_key}")
        return None

    data = []
    positions = []
    model_labels = []
    prompt_centers = []

    current_x = 1.0

    for prompt_type in PROMPT_TYPES:
        available = []

        for model_name in SAM_MODELS:
            if not PROMPT_SUPPORT[model_name][prompt_type]:
                continue

            subset = plot_df[
                (plot_df["Prompt type"] == prompt_type)
                & (plot_df["Model"] == model_name)
            ]["Value"].dropna().values

            if len(subset) > 0:
                available.append((model_name, subset))

        if len(available) == 0:
            current_x += GROUP_GAP
            continue

        offsets = get_offsets_for_models(len(available))

        for offset, (model_name, subset) in zip(offsets, available):
            position = current_x + offset
            data.append(subset)
            positions.append(position)
            model_labels.append(model_name)

        prompt_centers.append(
            {
                "prompt_type": prompt_type,
                "center": current_x,
            }
        )

        current_x += GROUP_GAP

    if len(data) == 0:
        print(f"No box data for plot: {dataset_key} | {split_key} | {metric_key}")
        return None

    fig, ax = plt.subplots(figsize=(FIG_WIDTH, FIG_HEIGHT))

    boxplot = ax.boxplot(
        data,
        positions=positions,
        widths=BOX_WIDTH,
        patch_artist=True,
        showmeans=True,
        meanline=True,
        showfliers=False,
        boxprops={
            "facecolor": BOX_COLOR,
            "edgecolor": "black",
            "linewidth": 1.1,
        },
        whiskerprops={
            "color": "black",
            "linewidth": 1.0,
        },
        capprops={
            "color": "black",
            "linewidth": 1.0,
        },
        medianprops={
            "color": MEDIAN_COLOR,
            "linewidth": 2.0,
        },
        meanprops={
            "color": MEAN_COLOR,
            "linewidth": 2.0,
            "linestyle": "-",
        },
    )

    for patch in boxplot["boxes"]:
        patch.set_facecolor(BOX_COLOR)
        patch.set_alpha(0.95)

    ax.set_title(
        f"{METRICS[metric_key]['title']} by Prompt Type and SAM Model\n{dataset_key} | {split_key}",
        fontsize=TITLE_FONT_SIZE,
        fontweight="bold",
        pad=10,
    )

    ax.set_ylabel(
        METRICS[metric_key]["ylabel"],
        fontsize=AXIS_FONT_SIZE,
        fontweight="bold",
    )

    ax.set_xticks([item["center"] for item in prompt_centers])
    ax.set_xticklabels(
        [item["prompt_type"] for item in prompt_centers],
        fontsize=TICK_FONT_SIZE,
        fontweight="bold",
    )

    # Push main prompt tick labels down slightly.
    ax.tick_params(axis="x", pad=8)

    y_min, y_max = ax.get_ylim()
    y_range = y_max - y_min

    # Put model labels below prompt tick labels, with more vertical separation.
    model_label_y = y_min - 0.155 * y_range

    for position, model_label in zip(positions, model_labels):
        ax.text(
            position,
            model_label_y,
            model_label,
            ha="right",
            va="top",
            rotation=45,
            fontsize=MODEL_LABEL_FONT_SIZE,
        )

    ax.set_ylim(
        bottom=max(0.0, y_min),
        top=min(1.05, y_max + 0.035 * y_range),
    )

    if len(prompt_centers) > 0:
        ax.set_xlim(
            prompt_centers[0]["center"] - 0.34,
            prompt_centers[-1]["center"] + 0.34,
        )

    ax.grid(axis="y", linestyle="--", alpha=0.35)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)

    legend_handles = [
        plt.Line2D(
            [0],
            [0],
            color=MEAN_COLOR,
            linewidth=2,
            label="Mean",
        ),
        plt.Line2D(
            [0],
            [0],
            color=MEDIAN_COLOR,
            linewidth=2,
            label="Median",
        ),
    ]

    ax.legend(
        handles=legend_handles,
        loc="lower right",
        fontsize=10,
        frameon=True,
    )

    fig.subplots_adjust(
        bottom=0.46,
        left=0.10,
        right=0.98,
        top=0.84,
    )

    FIGURE_DIR.mkdir(parents=True, exist_ok=True)

    output_png = FIGURE_DIR / f"{dataset_key}_{split_key}_{metric_key}_sam_prompt_boxplots_600dpi.png"

    fig.savefig(output_png, dpi=DPI, bbox_inches="tight")

    plt.close(fig)

    print(f"Saved PNG: {output_png}")

    return output_png


def make_all_boxplots(values_df):
    saved_rows = []

    for dataset_key in DATASETS:
        for split_key in SPLITS:
            for metric_key in METRICS.keys():
                output_png = make_single_boxplot(
                    values_df=values_df,
                    dataset_key=dataset_key,
                    split_key=split_key,
                    metric_key=metric_key,
                )

                if output_png is not None:
                    saved_rows.append(
                        {
                            "Dataset": dataset_key,
                            "Split": split_key,
                            "Metric": metric_key,
                            "PNG": str(output_png),
                        }
                    )

    return pd.DataFrame(saved_rows)


# =============================================================================
# Excel output
# =============================================================================

def autosize_columns(writer, sheet_name, df):
    worksheet = writer.sheets[sheet_name]

    for column_index, column_name in enumerate(df.columns, start=1):
        values = df[column_name].head(300).tolist()

        string_values = []
        for value in values:
            if pd.isna(value):
                string_values.append("")
            else:
                string_values.append(str(value))

        max_length = max([len(str(column_name))] + [len(v) for v in string_values])
        width = min(max(max_length + 2, 10), 90)

        worksheet.column_dimensions[get_column_letter(column_index)].width = width


def style_sheet(writer, sheet_name, df):
    worksheet = writer.sheets[sheet_name]
    worksheet.freeze_panes = "A2"

    header_fill = PatternFill(
        start_color="D9EAF7",
        end_color="D9EAF7",
        fill_type="solid",
    )

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    autosize_columns(writer, sheet_name, df)


def write_excel(
    prompt_support_df,
    simple_summary_df,
    wide_summary_df,
    saved_plots_df,
    values_df,
    discovered_df,
    check_df,
):
    OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(SUMMARY_XLSX, engine="openpyxl") as writer:
        simple_summary_df.to_excel(writer, sheet_name="simple_mean_std", index=False)
        wide_summary_df.to_excel(writer, sheet_name="wide_mean_std", index=False)
        prompt_support_df.to_excel(writer, sheet_name="prompt_support", index=False)
        saved_plots_df.to_excel(writer, sheet_name="saved_plots", index=False)
        values_df.to_excel(writer, sheet_name="image_level_metrics", index=False)
        discovered_df.to_excel(writer, sheet_name="discovered_files", index=False)
        check_df.to_excel(writer, sheet_name="metric_calculation_check", index=False)

        style_sheet(writer, "simple_mean_std", simple_summary_df)
        style_sheet(writer, "wide_mean_std", wide_summary_df)
        style_sheet(writer, "prompt_support", prompt_support_df)
        style_sheet(writer, "saved_plots", saved_plots_df)
        style_sheet(writer, "image_level_metrics", values_df)
        style_sheet(writer, "discovered_files", discovered_df)
        style_sheet(writer, "metric_calculation_check", check_df)

    print(f"Saved Excel: {SUMMARY_XLSX}")


# =============================================================================
# Main
# =============================================================================

def main():
    OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)
    FIGURE_DIR.mkdir(parents=True, exist_ok=True)

    print("Collecting SAM prompt comparison data...")

    discovered_df = discover_prompt_files()
    discovered_df.to_csv(DISCOVERED_CSV, index=False)
    print(f"Saved discovered file report: {DISCOVERED_CSV}")

    values_df, check_df = collect_all_metrics(discovered_df)
    values_df.to_csv(VALUES_CSV, index=False)
    check_df.to_csv(CHECK_CSV, index=False)

    print(f"Saved calculated metrics: {VALUES_CSV}")
    print(f"Saved calculation check: {CHECK_CSV}")
    print(f"Calculated image-level metric rows: {len(values_df)}")

    if len(values_df) == 0:
        print("\nERROR: No metrics were calculated.")
        print("Check these files:")
        print(DISCOVERED_CSV)
        print(CHECK_CSV)
        print("\nMost likely issue: prediction masks are not in the expected folder names.")
        return

    prompt_support_df = make_prompt_support_df()
    simple_summary_df = summarize_values(values_df)
    wide_summary_df = make_wide_summary(simple_summary_df)

    saved_plots_df = make_all_boxplots(values_df)

    print(f"Saved plot count: {len(saved_plots_df)}")

    write_excel(
        prompt_support_df=prompt_support_df,
        simple_summary_df=simple_summary_df,
        wide_summary_df=wide_summary_df,
        saved_plots_df=saved_plots_df,
        values_df=values_df,
        discovered_df=discovered_df,
        check_df=check_df,
    )

    print("\nDONE.")
    print(f"Figure folder: {FIGURE_DIR}")
    print(f"Summary Excel: {SUMMARY_XLSX}")


if __name__ == "__main__":
    main()