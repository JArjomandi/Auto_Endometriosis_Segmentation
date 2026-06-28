from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


# =============================================================================
# Paths
# =============================================================================

RESULTS_ROOT = Path(r"F:\Results\SAM_Benchmarking")

OUTPUT_DIR = (
    RESULTS_ROOT
    / "Model_comparison"
    / "inference_time_comparison"
)

OUTPUT_XLSX = OUTPUT_DIR / "all_models_mean_inference_times.xlsx"


# =============================================================================
# Datasets / splits
# =============================================================================

DATASETS = [
    "ENID",
    "GLENDA",
    "GLENDA_clean",
]

SPLITS = [
    "val",
    "test",
]


# =============================================================================
# Timing column names
# =============================================================================

TIME_COLUMN_PRIORITY = [
    "total_time_sec",
    "inference_time_sec",
    "inference_time_seconds",
    "prediction_time_sec",
    "predict_time_sec",
    "model_time_sec",
    "elapsed_time_sec",
    "runtime_sec",
    "duration_sec",
    "time_sec",
    "seconds_per_image",
    "total_seconds",
    "inference_seconds",
    "time_seconds",
    "inference_time_ms",
    "prediction_time_ms",
    "time_ms",
    "runtime_ms",
    "duration_ms",
]


BAD_TIME_KEYWORDS = [
    "train",
    "training",
    "epoch",
    "fit",
    "loss",
    "created",
    "modified",
    "timestamp",
    "date",
]


NON_TIMING_COLUMNS = {
    "dice",
    "iou",
    "precision",
    "recall",
    "sensitivity",
    "specificity",
    "accuracy",
    "f1",
    "tp",
    "fp",
    "fn",
    "tn",
    "pred_area",
    "gt_area",
    "num_boxes",
    "accepted_surgisam2_components",
    "fallback_to_segformer_components",
    "acceptance_rate_components",
}


# =============================================================================
# Discovery
# =============================================================================

def is_valid_result_file(path: Path):
    try:
        relative_parts = path.relative_to(RESULTS_ROOT).parts
    except Exception:
        return False

    if len(relative_parts) < 3:
        return False

    dataset = relative_parts[0]

    if dataset == "Model_comparison":
        return False

    if dataset not in DATASETS:
        return False

    if path.suffix.lower() not in [".csv", ".xlsx"]:
        return False

    # Avoid reading our own generated comparison files.
    if "Model_comparison" in relative_parts:
        return False

    return True


def find_all_result_tables():
    files = []

    for dataset in DATASETS:
        dataset_root = RESULTS_ROOT / dataset

        if not dataset_root.exists():
            print(f"WARNING: Missing dataset folder: {dataset_root}")
            continue

        for path in dataset_root.rglob("*"):
            if path.is_file() and is_valid_result_file(path):
                files.append(path)

    return sorted(files)


# =============================================================================
# Metadata parsing
# =============================================================================

def get_relative_parts(path: Path):
    return path.relative_to(RESULTS_ROOT).parts


def detect_dataset(path: Path):
    return get_relative_parts(path)[0]


def detect_model(path: Path):
    parts = get_relative_parts(path)

    if len(parts) >= 2:
        return parts[1]

    return "UNKNOWN_MODEL"


def detect_split_from_path(path: Path):
    parts = [part.lower() for part in get_relative_parts(path)]

    detected = None

    for part in parts:
        if part in SPLITS:
            detected = part

    return detected


def detect_split_from_dataframe(df: pd.DataFrame):
    split_columns = [
        "split",
        "Split",
        "dataset_split",
        "Dataset split",
        "phase",
        "Phase",
    ]

    for column in split_columns:
        if column not in df.columns:
            continue

        values = df[column].dropna().astype(str).str.lower().tolist()

        for value in values:
            for split in SPLITS:
                if value == split or split in value:
                    return split

    return None


def detect_split(path: Path, df: pd.DataFrame):
    split = detect_split_from_path(path)

    if split is not None:
        return split

    split = detect_split_from_dataframe(df)

    if split is not None:
        return split

    file_lower = path.name.lower()

    for split_name in SPLITS:
        if split_name in file_lower:
            return split_name

    return "unknown"


def detect_variant(path: Path, dataset: str, model: str, split: str):
    parts = list(get_relative_parts(path))

    # parts:
    # dataset / model / ... / file
    between_model_and_file = parts[2:-1]

    variant_parts = []

    for part in between_model_and_file:
        lower = part.lower()

        if lower in SPLITS:
            continue

        variant_parts.append(part)

    return "/".join(variant_parts)


def make_model_display(model: str, variant: str):
    if variant:
        return f"{model} | {variant}"
    return model


# =============================================================================
# File reading
# =============================================================================

def read_table_file(path: Path):
    """
    Returns list of (sheet_name, dataframe).
    CSV returns one dataframe.
    XLSX returns all sheets.
    """

    tables = []

    if path.suffix.lower() == ".csv":
        try:
            df = pd.read_csv(path)
            tables.append((path.name, df))
        except Exception as error:
            print(f"WARNING: Could not read CSV: {path} | {error}")

    elif path.suffix.lower() == ".xlsx":
        try:
            sheets = pd.read_excel(path, sheet_name=None)
            for sheet_name, df in sheets.items():
                tables.append((sheet_name, df))
        except Exception as error:
            print(f"WARNING: Could not read XLSX: {path} | {error}")

    return tables


# =============================================================================
# Timing column detection
# =============================================================================

def normalize_name(name):
    return str(name).strip().lower()


def column_name_suggests_timing(column_name):
    name = normalize_name(column_name)

    if name in NON_TIMING_COLUMNS:
        return False

    for bad in BAD_TIME_KEYWORDS:
        if bad in name:
            return False

    if name in TIME_COLUMN_PRIORITY:
        return True

    timing_keywords = [
        "inference",
        "predict",
        "prediction",
        "elapsed",
        "runtime",
        "duration",
        "time",
        "seconds",
        "_sec",
        "_ms",
    ]

    return any(keyword in name for keyword in timing_keywords)


def is_valid_timing_series(series):
    values = pd.to_numeric(series, errors="coerce").dropna()

    if len(values) == 0:
        return False

    if (values < 0).any():
        return False

    # Exclude obvious timestamps or total training durations.
    median_value = float(values.median())

    if median_value > 10000:
        return False

    return True


def find_timing_columns(df: pd.DataFrame):
    timing_columns = []

    for column in df.columns:
        if not column_name_suggests_timing(column):
            continue

        if not is_valid_timing_series(df[column]):
            continue

        timing_columns.append(column)

    return timing_columns


def choose_best_time_column(timing_columns):
    if len(timing_columns) == 0:
        return None

    lower_to_original = {
        normalize_name(column): column
        for column in timing_columns
    }

    for preferred_column in TIME_COLUMN_PRIORITY:
        if preferred_column in lower_to_original:
            return lower_to_original[preferred_column]

    # Prefer seconds over ms.
    second_columns = []

    for column in timing_columns:
        name = normalize_name(column)

        if (
            name.endswith("_sec")
            or "seconds" in name
            or name.endswith("_s")
        ):
            second_columns.append(column)

    if len(second_columns) > 0:
        return second_columns[0]

    return timing_columns[0]


def convert_to_seconds(values, column_name):
    values = pd.to_numeric(values, errors="coerce").dropna()
    name = normalize_name(column_name)

    if name.endswith("_ms") or "millisecond" in name:
        values = values / 1000.0

    return values


# =============================================================================
# Summaries
# =============================================================================

def summarize_values(values_sec):
    values_sec = pd.to_numeric(pd.Series(values_sec), errors="coerce").dropna()

    if len(values_sec) == 0:
        return None

    mean_value = float(values_sec.mean())
    std_value = float(values_sec.std(ddof=1)) if len(values_sec) > 1 else 0.0

    return {
        "N images": int(len(values_sec)),
        "Mean inference time (s)": mean_value,
        "Std inference time (s)": std_value,
        "Mean ± SD (s)": f"{mean_value:.4f} ± {std_value:.4f}",
        "Median inference time (s)": float(values_sec.median()),
        "Min inference time (s)": float(values_sec.min()),
        "Max inference time (s)": float(values_sec.max()),
        "Mean FPS": float(1.0 / mean_value) if mean_value > 0 else np.nan,
    }


def analyze_one_table(path: Path, sheet_name: str, df: pd.DataFrame):
    if df is None or len(df) == 0:
        return None

    timing_columns = find_timing_columns(df)

    if len(timing_columns) == 0:
        return None

    time_column = choose_best_time_column(timing_columns)

    if time_column is None:
        return None

    values_sec = convert_to_seconds(
        values=df[time_column],
        column_name=time_column,
    )

    summary = summarize_values(values_sec)

    if summary is None:
        return None

    dataset = detect_dataset(path)
    model = detect_model(path)
    split = detect_split(path, df)

    variant = detect_variant(
        path=path,
        dataset=dataset,
        model=model,
        split=split,
    )

    model_display = make_model_display(
        model=model,
        variant=variant,
    )

    row = {
        "Dataset": dataset,
        "Split": split,
        "Model": model,
        "Variant": variant,
        "Model display": model_display,
        "Time column used": time_column,
        "All timing columns found": ", ".join([str(col) for col in timing_columns]),
        "Source file": path.name,
        "Source sheet": sheet_name,
        "Source path": str(path),
    }

    row.update(summary)

    return row


def collect_all_summaries():
    files = find_all_result_tables()

    print(f"Result CSV/XLSX files found: {len(files)}")

    timing_rows = []
    no_timing_rows = []

    for path in files:
        tables = read_table_file(path)

        if len(tables) == 0:
            continue

        file_has_timing = False

        for sheet_name, df in tables:
            row = analyze_one_table(
                path=path,
                sheet_name=sheet_name,
                df=df,
            )

            if row is not None:
                timing_rows.append(row)
                file_has_timing = True

        if not file_has_timing:
            dataset = detect_dataset(path)
            model = detect_model(path)
            split = detect_split_from_path(path)

            no_timing_rows.append(
                {
                    "Dataset": dataset,
                    "Split": split if split is not None else "unknown",
                    "Model": model,
                    "File": path.name,
                    "Path": str(path),
                    "Note": "No usable timing column detected.",
                }
            )

    timing_df = pd.DataFrame(timing_rows)
    no_timing_df = pd.DataFrame(no_timing_rows)

    return timing_df, no_timing_df


# =============================================================================
# Deduplication
# =============================================================================

def source_priority(row):
    """
    Lower is better.

    Prefer:
        1. inference_results.csv for SAM GT_box and similar SAM outputs
        2. inference_times.csv
        3. metrics_image_level.csv
        4. anything else
    """

    source_file = str(row["Source file"]).lower()
    source_path = str(row["Source path"]).lower()

    if source_file == "inference_results.csv" and "gt_box" in source_path:
        return 1

    if source_file == "inference_times.csv":
        return 2

    if source_file == "inference_results.csv":
        return 3

    if source_file == "metrics_image_level.csv":
        return 4

    return 5


def deduplicate_rows(timing_df: pd.DataFrame):
    """
    Keeps one row per:
        Dataset + Split + Model display

    This avoids duplicates when several CSV files contain timing.
    """

    if len(timing_df) == 0:
        return timing_df

    df = timing_df.copy()
    df["source_priority"] = df.apply(source_priority, axis=1)

    df = df.sort_values(
        by=[
            "Dataset",
            "Split",
            "Model display",
            "source_priority",
            "N images",
        ],
        ascending=[
            True,
            True,
            True,
            True,
            False,
        ],
    )

    dedup_df = df.drop_duplicates(
        subset=[
            "Dataset",
            "Split",
            "Model display",
        ],
        keep="first",
    ).copy()

    dedup_df = dedup_df.drop(columns=["source_priority"])

    return dedup_df


# =============================================================================
# Excel output
# =============================================================================

def make_compact_table(dedup_df: pd.DataFrame):
    columns = [
        "Dataset",
        "Split",
        "Model display",
        "N images",
        "Mean ± SD (s)",
        "Mean inference time (s)",
        "Std inference time (s)",
        "Median inference time (s)",
        "Mean FPS",
        "Time column used",
    ]

    existing = [
        column for column in columns
        if column in dedup_df.columns
    ]

    return dedup_df[existing].copy()


def make_wide_table(compact_df: pd.DataFrame):
    if len(compact_df) == 0:
        return pd.DataFrame()

    wide_df = compact_df.pivot_table(
        index="Model display",
        columns=["Dataset", "Split"],
        values="Mean ± SD (s)",
        aggfunc="first",
    ).reset_index()

    wide_df.columns = [
        " ".join([str(part) for part in column if str(part) != ""]).strip()
        if isinstance(column, tuple)
        else str(column)
        for column in wide_df.columns
    ]

    return wide_df


def autosize_columns(writer, sheet_name: str, df: pd.DataFrame):
    worksheet = writer.sheets[sheet_name]

    for column_index, column_name in enumerate(df.columns, start=1):
        raw_values = df[column_name].head(300).tolist()

        string_values = []

        for value in raw_values:
            if pd.isna(value):
                string_values.append("")
            else:
                string_values.append(str(value))

        max_length = max(
            [len(str(column_name))]
            + [len(value) for value in string_values]
        )

        width = min(max(max_length + 2, 10), 100)

        column_letter = get_column_letter(column_index)
        worksheet.column_dimensions[column_letter].width = width


def style_sheet(writer, sheet_name: str, df: pd.DataFrame):
    worksheet = writer.sheets[sheet_name]
    worksheet.freeze_panes = "A2"

    header_fill = PatternFill(
        start_color="D9EAF7",
        end_color="D9EAF7",
        fill_type="solid",
    )

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    autosize_columns(writer, sheet_name, df)


def write_excel(
    compact_df: pd.DataFrame,
    wide_df: pd.DataFrame,
    dedup_df: pd.DataFrame,
    all_detected_df: pd.DataFrame,
    no_timing_df: pd.DataFrame,
):
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        compact_df.to_excel(writer, sheet_name="mean_std_compact", index=False)
        wide_df.to_excel(writer, sheet_name="wide_table", index=False)
        dedup_df.to_excel(writer, sheet_name="deduplicated_details", index=False)
        all_detected_df.to_excel(writer, sheet_name="all_detected_timing", index=False)

        if len(no_timing_df) > 0:
            no_timing_df.to_excel(writer, sheet_name="files_without_timing", index=False)

        style_sheet(writer, "mean_std_compact", compact_df)
        style_sheet(writer, "wide_table", wide_df)
        style_sheet(writer, "deduplicated_details", dedup_df)
        style_sheet(writer, "all_detected_timing", all_detected_df)

        if len(no_timing_df) > 0:
            style_sheet(writer, "files_without_timing", no_timing_df)

    print(f"Saved Excel: {OUTPUT_XLSX}")


# =============================================================================
# Main
# =============================================================================

def main():
    timing_df, no_timing_df = collect_all_summaries()

    if len(timing_df) == 0:
        raise RuntimeError(
            "No inference-time columns were found in any CSV/XLSX files. "
            "This means the times may not have been saved for the non-SAM models."
        )

    timing_df = timing_df.sort_values(
        by=[
            "Dataset",
            "Split",
            "Model display",
            "Source file",
        ]
    )

    dedup_df = deduplicate_rows(timing_df)

    dedup_df = dedup_df.sort_values(
        by=[
            "Dataset",
            "Split",
            "Model display",
        ]
    )

    compact_df = make_compact_table(dedup_df)
    wide_df = make_wide_table(compact_df)

    write_excel(
        compact_df=compact_df,
        wide_df=wide_df,
        dedup_df=dedup_df,
        all_detected_df=timing_df,
        no_timing_df=no_timing_df,
    )

    print("\nDONE.")
    print(f"Output: {OUTPUT_XLSX}")
    print("\nUse sheet: mean_std_compact")
    print("If a model is missing, check sheet: files_without_timing")


if __name__ == "__main__":
    main()