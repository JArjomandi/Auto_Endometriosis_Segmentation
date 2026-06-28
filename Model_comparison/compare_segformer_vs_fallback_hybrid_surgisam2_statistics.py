from pathlib import Path
import math
import json

import numpy as np
import pandas as pd
from scipy.stats import wilcoxon, norm
from openpyxl.styles import PatternFill, Font


# =============================================================================
# Paths
# =============================================================================

RESULTS_ROOT = Path(r"F:\Results\SAM_Benchmarking")

OUTPUT_DIR = (
    RESULTS_ROOT
    / "Model_comparison"
    / "statistical_comparison"
)

OUTPUT_XLSX = OUTPUT_DIR / "segformer_vs_fallback_hybrid_surgisam2_statistics.xlsx"


# =============================================================================
# Fallback hybrid result folder settings
# =============================================================================

HYBRID_METHOD_FOLDER = "SegFormer_SurgiSAM2_AutoBox_Fallback"
HYBRID_TRAINING_STATE = "hybrid"

# This must match the tuned fallback runner output folder.
HYBRID_PROMPT_MODE = "Auto_box_fallback_dice_0p85_area_0p70_1p30"


DATASETS = [
    "ENID",
    "GLENDA",
    "GLENDA_clean",
]

SPLITS = [
    "val",
    "test",
]


# SegFormer columns = before SurgiSAM2 fallback refinement.
# Hybrid columns = after tuned fallback hybrid refinement.
METRIC_PAIRS = {
    "dice": {
        "segformer_col": "segformer_initial_dice",
        "hybrid_col": "dice",
    },
    "iou": {
        "segformer_col": "segformer_initial_iou",
        "hybrid_col": "iou",
    },
    "precision": {
        "segformer_col": "segformer_initial_precision",
        "hybrid_col": "precision",
    },
    "recall": {
        "segformer_col": "segformer_initial_recall",
        "hybrid_col": "recall",
    },
}


ALPHA = 0.05


# =============================================================================
# Path helpers
# =============================================================================

def get_hybrid_metrics_path(dataset_key: str, split_key: str):
    return (
        RESULTS_ROOT
        / dataset_key
        / HYBRID_METHOD_FOLDER
        / HYBRID_TRAINING_STATE
        / HYBRID_PROMPT_MODE
        / split_key
        / "metrics_image_level.csv"
    )


# =============================================================================
# Numeric helpers
# =============================================================================

def safe_float_series(series):
    return pd.to_numeric(series, errors="coerce")


def summarize_values(values):
    values = pd.to_numeric(pd.Series(values), errors="coerce").dropna()

    if len(values) == 0:
        return {
            "n": 0,
            "mean": np.nan,
            "std": np.nan,
            "median": np.nan,
            "q1": np.nan,
            "q3": np.nan,
            "min": np.nan,
            "max": np.nan,
        }

    return {
        "n": int(len(values)),
        "mean": float(values.mean()),
        "std": float(values.std(ddof=1)) if len(values) > 1 else 0.0,
        "median": float(values.median()),
        "q1": float(values.quantile(0.25)),
        "q3": float(values.quantile(0.75)),
        "min": float(values.min()),
        "max": float(values.max()),
    }


def format_mean_sd(mean_value, sd_value):
    if not np.isfinite(mean_value):
        return ""

    if not np.isfinite(sd_value):
        return f"{mean_value:.4f}"

    return f"{mean_value:.4f} ± {sd_value:.4f}"


def format_median_iqr(median_value, q1_value, q3_value):
    if not np.isfinite(median_value):
        return ""

    if not np.isfinite(q1_value) or not np.isfinite(q3_value):
        return f"{median_value:.4f}"

    return f"{median_value:.4f} [{q1_value:.4f}, {q3_value:.4f}]"


# =============================================================================
# Statistical tests
# =============================================================================

def run_wilcoxon_paired(segformer_values, hybrid_values):
    """
    Paired Wilcoxon signed-rank test.

    Direction:
        difference = hybrid - segformer

    Positive difference:
        tuned fallback hybrid improved over SegFormer.

    Negative difference:
        tuned fallback hybrid worsened compared with SegFormer.
    """

    segformer_values = np.asarray(segformer_values, dtype=np.float64)
    hybrid_values = np.asarray(hybrid_values, dtype=np.float64)

    valid_mask = np.isfinite(segformer_values) & np.isfinite(hybrid_values)

    segformer_values = segformer_values[valid_mask]
    hybrid_values = hybrid_values[valid_mask]

    differences = hybrid_values - segformer_values
    nonzero_differences = differences[differences != 0]

    if len(segformer_values) == 0:
        return {
            "n_paired": 0,
            "n_nonzero_differences": 0,
            "wilcoxon_statistic": np.nan,
            "p_value": np.nan,
            "z_approx_abs": np.nan,
            "effect_size_r_abs": np.nan,
            "test_note": "No valid paired values.",
        }

    if len(nonzero_differences) == 0:
        return {
            "n_paired": int(len(segformer_values)),
            "n_nonzero_differences": 0,
            "wilcoxon_statistic": 0.0,
            "p_value": 1.0,
            "z_approx_abs": 0.0,
            "effect_size_r_abs": 0.0,
            "test_note": "All paired differences are zero.",
        }

    try:
        result = wilcoxon(
            hybrid_values,
            segformer_values,
            alternative="two-sided",
            zero_method="wilcox",
            correction=False,
            mode="auto",
        )

        p_value = float(result.pvalue)
        statistic = float(result.statistic)

    except Exception as error:
        return {
            "n_paired": int(len(segformer_values)),
            "n_nonzero_differences": int(len(nonzero_differences)),
            "wilcoxon_statistic": np.nan,
            "p_value": np.nan,
            "z_approx_abs": np.nan,
            "effect_size_r_abs": np.nan,
            "test_note": f"Wilcoxon failed: {error}",
        }

    if p_value <= 0:
        z_abs = np.inf
    elif p_value >= 1:
        z_abs = 0.0
    else:
        z_abs = float(norm.isf(p_value / 2.0))

    if len(nonzero_differences) > 0 and np.isfinite(z_abs):
        effect_size_r_abs = float(z_abs / math.sqrt(len(nonzero_differences)))
    else:
        effect_size_r_abs = np.nan

    return {
        "n_paired": int(len(segformer_values)),
        "n_nonzero_differences": int(len(nonzero_differences)),
        "wilcoxon_statistic": statistic,
        "p_value": p_value,
        "z_approx_abs": z_abs,
        "effect_size_r_abs": effect_size_r_abs,
        "test_note": "OK",
    }


def holm_adjust_pvalues(p_values):
    """
    Holm-Bonferroni correction.

    Input:
        list/array of raw p-values. May contain NaN.

    Output:
        adjusted p-values in original order.
    """

    p_values = np.asarray(p_values, dtype=np.float64)
    adjusted = np.full_like(p_values, np.nan, dtype=np.float64)

    valid_indices = np.where(np.isfinite(p_values))[0]

    if len(valid_indices) == 0:
        return adjusted

    valid_p = p_values[valid_indices]
    order = np.argsort(valid_p)

    sorted_indices = valid_indices[order]
    sorted_p = valid_p[order]

    m = len(sorted_p)
    adjusted_sorted = np.zeros(m, dtype=np.float64)

    running_max = 0.0

    for rank_index, p_value in enumerate(sorted_p):
        multiplier = m - rank_index
        adjusted_p = min(1.0, p_value * multiplier)
        running_max = max(running_max, adjusted_p)
        adjusted_sorted[rank_index] = running_max

    for sorted_position, original_index in enumerate(sorted_indices):
        adjusted[original_index] = adjusted_sorted[sorted_position]

    return adjusted


def decide_better(median_difference, mean_difference, adjusted_p):
    """
    Positive difference:
        Fallback hybrid > SegFormer

    Negative difference:
        SegFormer > fallback hybrid
    """

    if not np.isfinite(adjusted_p):
        return "Unable to test"

    if adjusted_p >= ALPHA:
        if median_difference > 0:
            return "Fallback hybrid numerically higher, not significant"
        if median_difference < 0:
            return "SegFormer numerically higher, not significant"
        return "No difference"

    if median_difference > 0:
        return "Fallback hybrid significantly better"

    if median_difference < 0:
        return "SegFormer significantly better"

    if mean_difference > 0:
        return "Fallback hybrid significantly better by mean difference"

    if mean_difference < 0:
        return "SegFormer significantly better by mean difference"

    return "Significant test but zero median/mean difference"


# =============================================================================
# Main analysis
# =============================================================================

def extract_optional_summary_columns(df: pd.DataFrame):
    """
    Pulls optional fallback-specific columns if they exist.
    These are not used in the statistical test, but are useful for interpretation.
    """

    optional_columns = [
        "num_boxes",
        "accepted_surgisam2_components",
        "fallback_to_segformer_components",
        "acceptance_rate_components",
        "acceptance_iou_threshold",
        "acceptance_dice_threshold",
        "min_area_ratio",
        "max_area_ratio",
        "dilation_radius_px",
    ]

    output = {}

    for column in optional_columns:
        if column not in df.columns:
            output[f"{column} mean"] = np.nan
            output[f"{column} median"] = np.nan
            continue

        values = pd.to_numeric(df[column], errors="coerce").dropna()

        if len(values) == 0:
            output[f"{column} mean"] = np.nan
            output[f"{column} median"] = np.nan
        else:
            output[f"{column} mean"] = float(values.mean())
            output[f"{column} median"] = float(values.median())

    return output


def analyze_one_dataset_split(dataset_key: str, split_key: str):
    metrics_path = get_hybrid_metrics_path(dataset_key, split_key)

    if not metrics_path.exists():
        print(f"WARNING: Missing metrics file: {metrics_path}")
        return [], []

    df = pd.read_csv(metrics_path)

    fallback_optional_summary = extract_optional_summary_columns(df)

    summary_rows = []
    image_level_rows = []

    for metric_name, pair in METRIC_PAIRS.items():
        seg_col = pair["segformer_col"]
        hybrid_col = pair["hybrid_col"]

        if seg_col not in df.columns or hybrid_col not in df.columns:
            print(
                f"WARNING: Missing columns for {dataset_key} | {split_key} | {metric_name}: "
                f"{seg_col}, {hybrid_col}"
            )
            continue

        segformer_values = safe_float_series(df[seg_col])
        hybrid_values = safe_float_series(df[hybrid_col])

        valid_mask = segformer_values.notna() & hybrid_values.notna()

        segformer_values_valid = segformer_values[valid_mask]
        hybrid_values_valid = hybrid_values[valid_mask]
        differences = hybrid_values_valid - segformer_values_valid

        seg_summary = summarize_values(segformer_values_valid)
        hybrid_summary = summarize_values(hybrid_values_valid)
        diff_summary = summarize_values(differences)

        test_result = run_wilcoxon_paired(
            segformer_values=segformer_values_valid.values,
            hybrid_values=hybrid_values_valid.values,
        )

        improved_count = int((differences > 0).sum())
        worsened_count = int((differences < 0).sum())
        unchanged_count = int((differences == 0).sum())

        n_valid = int(len(differences))

        improvement_rate = improved_count / n_valid if n_valid > 0 else np.nan
        worsening_rate = worsened_count / n_valid if n_valid > 0 else np.nan
        unchanged_rate = unchanged_count / n_valid if n_valid > 0 else np.nan

        row = {
            "Dataset": dataset_key,
            "Split": split_key,
            "Metric": metric_name,

            "n paired": test_result["n_paired"],
            "n nonzero differences": test_result["n_nonzero_differences"],

            "SegFormer mean": seg_summary["mean"],
            "SegFormer SD": seg_summary["std"],
            "SegFormer mean ± SD": format_mean_sd(
                seg_summary["mean"],
                seg_summary["std"],
            ),
            "SegFormer median": seg_summary["median"],
            "SegFormer Q1": seg_summary["q1"],
            "SegFormer Q3": seg_summary["q3"],
            "SegFormer median [Q1, Q3]": format_median_iqr(
                seg_summary["median"],
                seg_summary["q1"],
                seg_summary["q3"],
            ),

            "Fallback hybrid mean": hybrid_summary["mean"],
            "Fallback hybrid SD": hybrid_summary["std"],
            "Fallback hybrid mean ± SD": format_mean_sd(
                hybrid_summary["mean"],
                hybrid_summary["std"],
            ),
            "Fallback hybrid median": hybrid_summary["median"],
            "Fallback hybrid Q1": hybrid_summary["q1"],
            "Fallback hybrid Q3": hybrid_summary["q3"],
            "Fallback hybrid median [Q1, Q3]": format_median_iqr(
                hybrid_summary["median"],
                hybrid_summary["q1"],
                hybrid_summary["q3"],
            ),

            "Mean difference FallbackHybrid-SegFormer": diff_summary["mean"],
            "SD difference": diff_summary["std"],
            "Median difference FallbackHybrid-SegFormer": diff_summary["median"],
            "Q1 difference": diff_summary["q1"],
            "Q3 difference": diff_summary["q3"],
            "Min difference": diff_summary["min"],
            "Max difference": diff_summary["max"],

            "Improved images": improved_count,
            "Worsened images": worsened_count,
            "Unchanged images": unchanged_count,
            "Improvement rate": improvement_rate,
            "Worsening rate": worsening_rate,
            "Unchanged rate": unchanged_rate,

            "Wilcoxon statistic": test_result["wilcoxon_statistic"],
            "Wilcoxon p": test_result["p_value"],
            "Holm-adjusted p": np.nan,
            "Significant after Holm": "",
            "Effect size r abs": test_result["effect_size_r_abs"],
            "z approx abs": test_result["z_approx_abs"],
            "Better method": "",
            "Test note": test_result["test_note"],
            "Metrics file": str(metrics_path),
        }

        row.update(fallback_optional_summary)
        summary_rows.append(row)

        for index in df[valid_mask].index:
            image_name = df.loc[index, "image_name"] if "image_name" in df.columns else ""
            case_id = df.loc[index, "case_id"] if "case_id" in df.columns else ""

            segformer_value = float(segformer_values.loc[index])
            hybrid_value = float(hybrid_values.loc[index])
            difference_value = hybrid_value - segformer_value

            image_level_row = {
                "Dataset": dataset_key,
                "Split": split_key,
                "Metric": metric_name,
                "case_id": case_id,
                "image_name": image_name,
                "SegFormer": segformer_value,
                "Fallback hybrid": hybrid_value,
                "Difference FallbackHybrid-SegFormer": difference_value,
                "Improved": "Yes" if difference_value > 0 else "No",
                "Worsened": "Yes" if difference_value < 0 else "No",
                "Unchanged": "Yes" if difference_value == 0 else "No",
            }

            optional_image_columns = [
                "num_boxes",
                "accepted_surgisam2_components",
                "fallback_to_segformer_components",
                "acceptance_rate_components",
                "acceptance_iou_threshold",
                "acceptance_dice_threshold",
                "min_area_ratio",
                "max_area_ratio",
                "dilation_radius_px",
            ]

            for optional_column in optional_image_columns:
                if optional_column in df.columns:
                    image_level_row[optional_column] = df.loc[index, optional_column]

            image_level_rows.append(image_level_row)

    return summary_rows, image_level_rows


def apply_holm_within_dataset_split(summary_df: pd.DataFrame):
    """
    Holm correction is applied within each Dataset + Split across the tested metrics.

    Example:
        ENID test has 4 p-values:
            Dice, IoU, precision, recall

    These 4 p-values are Holm-corrected together.
    """

    summary_df = summary_df.copy()

    summary_df["Holm-adjusted p"] = np.nan
    summary_df["Significant after Holm"] = ""
    summary_df["Better method"] = ""

    for (dataset_key, split_key), group_df in summary_df.groupby(["Dataset", "Split"]):
        group_indices = group_df.index.to_list()
        p_values = group_df["Wilcoxon p"].values

        adjusted_values = holm_adjust_pvalues(p_values)

        for index, adjusted_p in zip(group_indices, adjusted_values):
            summary_df.loc[index, "Holm-adjusted p"] = adjusted_p

            significant = (
                np.isfinite(adjusted_p)
                and adjusted_p < ALPHA
            )

            summary_df.loc[index, "Significant after Holm"] = (
                "Yes" if significant else "No"
            )

            median_difference = summary_df.loc[
                index,
                "Median difference FallbackHybrid-SegFormer",
            ]

            mean_difference = summary_df.loc[
                index,
                "Mean difference FallbackHybrid-SegFormer",
            ]

            summary_df.loc[index, "Better method"] = decide_better(
                median_difference=median_difference,
                mean_difference=mean_difference,
                adjusted_p=adjusted_p,
            )

    return summary_df


# =============================================================================
# Interpretation/config sheets
# =============================================================================

def create_interpretation_sheet(summary_df: pd.DataFrame):
    rows = []

    rows.append(
        {
            "Item": "Comparison",
            "Description": (
                "SegFormer initial prediction versus SegFormer-guided SurgiSAM2 "
                "tuned fallback hybrid final prediction."
            ),
        }
    )

    rows.append(
        {
            "Item": "Fallback method",
            "Description": (
                "For each SegFormer connected component, SurgiSAM2 candidates are evaluated. "
                "A candidate is accepted only if it passes the tuned agreement criteria; "
                "otherwise the original SegFormer component is retained."
            ),
        }
    )

    rows.append(
        {
            "Item": "Pairing",
            "Description": (
                "Paired by image. Each image has one SegFormer value and one fallback hybrid "
                "value for the same metric."
            ),
        }
    )

    rows.append(
        {
            "Item": "Difference",
            "Description": (
                "Difference = Fallback hybrid - SegFormer. Positive values mean the fallback "
                "hybrid improved the metric. Negative values mean fallback refinement worsened it."
            ),
        }
    )

    rows.append(
        {
            "Item": "Statistical test",
            "Description": (
                "Two-sided Wilcoxon signed-rank test, because the two methods are paired "
                "on the same images and segmentation metrics are bounded/skewed."
            ),
        }
    )

    rows.append(
        {
            "Item": "Multiple testing correction",
            "Description": (
                "Holm correction is applied within each dataset and split across "
                "Dice, IoU, precision, and recall."
            ),
        }
    )

    rows.append(
        {
            "Item": "Decision rule",
            "Description": (
                "Fallback hybrid is significantly better if median difference > 0 and "
                "Holm-adjusted p < 0.05. SegFormer is significantly better if "
                "median difference < 0 and Holm-adjusted p < 0.05."
            ),
        }
    )

    rows.append(
        {
            "Item": "Primary metric",
            "Description": (
                "Dice should be treated as the primary metric. IoU, precision, "
                "and recall are secondary."
            ),
        }
    )

    rows.append(
        {
            "Item": "How to report",
            "Description": (
                "Report mean ± SD, median [Q1, Q3], median difference, "
                "Holm-adjusted p-value, effect size r, and Better method."
            ),
        }
    )

    if len(summary_df) > 0:
        dice_rows = summary_df[summary_df["Metric"] == "dice"].copy()

        for _, row in dice_rows.iterrows():
            rows.append(
                {
                    "Item": f"{row['Dataset']} {row['Split']} Dice conclusion",
                    "Description": (
                        f"{row['Better method']} | "
                        f"SegFormer mean={row['SegFormer mean']:.4f}, "
                        f"Fallback hybrid mean={row['Fallback hybrid mean']:.4f}, "
                        f"median difference={row['Median difference FallbackHybrid-SegFormer']:.4f}, "
                        f"Holm-adjusted p={row['Holm-adjusted p']:.4g}."
                    ),
                }
            )

    return pd.DataFrame(rows)


def create_config_sheet():
    return pd.DataFrame(
        [
            {
                "Parameter": "RESULTS_ROOT",
                "Value": str(RESULTS_ROOT),
            },
            {
                "Parameter": "HYBRID_METHOD_FOLDER",
                "Value": HYBRID_METHOD_FOLDER,
            },
            {
                "Parameter": "HYBRID_TRAINING_STATE",
                "Value": HYBRID_TRAINING_STATE,
            },
            {
                "Parameter": "HYBRID_PROMPT_MODE",
                "Value": HYBRID_PROMPT_MODE,
            },
            {
                "Parameter": "DATASETS",
                "Value": json.dumps(DATASETS),
            },
            {
                "Parameter": "SPLITS",
                "Value": json.dumps(SPLITS),
            },
            {
                "Parameter": "METRICS",
                "Value": json.dumps(list(METRIC_PAIRS.keys())),
            },
            {
                "Parameter": "ALPHA",
                "Value": ALPHA,
            },
            {
                "Parameter": "OUTPUT_XLSX",
                "Value": str(OUTPUT_XLSX),
            },
        ]
    )


# =============================================================================
# Excel writing
# =============================================================================

def autosize_worksheet_columns(writer, sheet_name: str, df: pd.DataFrame, max_width: int = 70):
    worksheet = writer.sheets[sheet_name]

    for column_index, column_name in enumerate(df.columns, start=1):
        column_values = df[column_name].astype(str).replace("nan", "")
        max_len = max(
            [len(str(column_name))]
            + [len(value) for value in column_values.head(200)]
        )

        width = min(max(max_len + 2, 10), max_width)
        column_letter = worksheet.cell(row=1, column=column_index).column_letter
        worksheet.column_dimensions[column_letter].width = width


def style_excel_sheet(writer, sheet_name: str, df: pd.DataFrame):
    worksheet = writer.sheets[sheet_name]
    worksheet.freeze_panes = "A2"

    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    autosize_worksheet_columns(writer, sheet_name, df)


def write_xlsx(summary_df, image_level_df, interpretation_df, config_df):
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        interpretation_df.to_excel(writer, sheet_name="interpretation", index=False)
        summary_df.to_excel(writer, sheet_name="summary_all", index=False)
        image_level_df.to_excel(writer, sheet_name="image_level_all", index=False)
        config_df.to_excel(writer, sheet_name="run_config", index=False)

        style_excel_sheet(writer, "interpretation", interpretation_df)
        style_excel_sheet(writer, "summary_all", summary_df)
        style_excel_sheet(writer, "image_level_all", image_level_df)
        style_excel_sheet(writer, "run_config", config_df)

        summary_sheet = writer.sheets["summary_all"]

        better_method_col = summary_df.columns.get_loc("Better method") + 1

        green_fill = PatternFill(
            start_color="C6EFCE",
            end_color="C6EFCE",
            fill_type="solid",
        )

        red_fill = PatternFill(
            start_color="FFC7CE",
            end_color="FFC7CE",
            fill_type="solid",
        )

        yellow_fill = PatternFill(
            start_color="FFEB9C",
            end_color="FFEB9C",
            fill_type="solid",
        )

        for row in range(2, summary_sheet.max_row + 1):
            better_method = summary_sheet.cell(
                row=row,
                column=better_method_col,
            ).value

            target_cell = summary_sheet.cell(
                row=row,
                column=better_method_col,
            )

            if better_method == "Fallback hybrid significantly better":
                target_cell.fill = green_fill

            elif better_method == "SegFormer significantly better":
                target_cell.fill = red_fill

            elif isinstance(better_method, str) and "not significant" in better_method:
                target_cell.fill = yellow_fill

    print(f"Saved XLSX: {OUTPUT_XLSX}")


# =============================================================================
# Main
# =============================================================================

def main():
    all_summary_rows = []
    all_image_level_rows = []

    for dataset_key in DATASETS:
        for split_key in SPLITS:
            print(f"Analyzing: {dataset_key} | {split_key}")

            summary_rows, image_level_rows = analyze_one_dataset_split(
                dataset_key=dataset_key,
                split_key=split_key,
            )

            all_summary_rows.extend(summary_rows)
            all_image_level_rows.extend(image_level_rows)

    summary_df = pd.DataFrame(all_summary_rows)

    if len(summary_df) == 0:
        raise RuntimeError(
            "No summary rows were created. Check that fallback hybrid metrics_image_level.csv files exist."
        )

    summary_df = apply_holm_within_dataset_split(summary_df)

    image_level_df = pd.DataFrame(all_image_level_rows)

    interpretation_df = create_interpretation_sheet(summary_df)
    config_df = create_config_sheet()

    write_xlsx(
        summary_df=summary_df,
        image_level_df=image_level_df,
        interpretation_df=interpretation_df,
        config_df=config_df,
    )

    print("\nDONE.")
    print(f"Output: {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()