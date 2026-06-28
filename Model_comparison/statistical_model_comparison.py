from pathlib import Path
import sys

import numpy as np
import pandas as pd
from scipy.stats import wilcoxon, norm


PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))


RESULTS_ROOT = Path(r"F:\Results\SAM_Benchmarking")
OUTPUT_ROOT = RESULTS_ROOT / "Model_comparison" / "statistical_comparison"


OUTPUT_XLSX = OUTPUT_ROOT / "simple_model_statistics_tables.xlsx"


DATASETS = [
    "ENID",
    "GLENDA",
    "GLENDA_clean",
]


SPLITS = [
    "val",
    "test",
]


METRICS = [
    "dice",
    "iou",
    "recall",
    "precision",
]


SAM_PROMPT_MODE_TO_COMPARE = None

SAM_PROMPT_MODE_CANDIDATES = [
    "GT_box",
    "GT_box_point",
    "GT_point",
    "Box_prompt",
    "Point_prompt",
    "No_prompt",
    "Auto_YOLO_box",
]


MODEL_SPECS = [
    {
        "display_name": "SAM2",
        "folder_candidates": ["SAM2"],
        "training_state": "frozen",
        "prompt_mode": SAM_PROMPT_MODE_TO_COMPARE,
        "is_sam": True,
        "group": "SAM",
    },
    {
        "display_name": "MedSAM",
        "folder_candidates": ["MedSAM"],
        "training_state": "frozen",
        "prompt_mode": SAM_PROMPT_MODE_TO_COMPARE,
        "is_sam": True,
        "group": "SAM",
    },
    {
        "display_name": "SAM-Med2D",
        "folder_candidates": ["SAM-Med2D", "SAMMed2D", "SAM_Med2D"],
        "training_state": "frozen",
        "prompt_mode": SAM_PROMPT_MODE_TO_COMPARE,
        "is_sam": True,
        "group": "SAM",
    },
    {
        "display_name": "SurgiSAM2",
        "folder_candidates": ["SurgiSAM2"],
        "training_state": "frozen",
        "prompt_mode": SAM_PROMPT_MODE_TO_COMPARE,
        "is_sam": True,
        "group": "SAM",
    },
    {
        "display_name": "YOLO11s-seg",
        "folder_candidates": ["YOLO11s_seg"],
        "training_state": "trained",
        "prompt_mode": "No_prompt",
        "is_sam": False,
        "group": "Trained",
    },
    {
        "display_name": "DeepLabV3+",
        "folder_candidates": ["DeepLabV3Plus"],
        "training_state": "trained",
        "prompt_mode": "No_prompt",
        "is_sam": False,
        "group": "Trained",
    },
    {
        "display_name": "SegFormer",
        "folder_candidates": ["SegFormer"],
        "training_state": "trained",
        "prompt_mode": "No_prompt",
        "is_sam": False,
        "group": "Trained",
    },
    {
        "display_name": "UNet++",
        "folder_candidates": ["UNetPP"],
        "training_state": "trained",
        "prompt_mode": "No_prompt",
        "is_sam": False,
        "group": "Trained",
    },
    {
        "display_name": "nnU-Net v2 2D",
        "folder_candidates": ["nnUNetV2_2D_100ep", "nnUNetV2_2D"],
        "training_state": "trained",
        "prompt_mode": "No_prompt",
        "is_sam": False,
        "group": "Trained",
    },
]


GROUPS = [
    "SAM",
    "Trained",
]


def find_metrics_file_for_model(dataset_key: str, split_key: str, model_spec: dict):
    for folder_name in model_spec["folder_candidates"]:
        model_root = (
            RESULTS_ROOT
            / dataset_key
            / folder_name
            / model_spec["training_state"]
        )

        if not model_root.exists():
            continue

        if model_spec["is_sam"]:
            if model_spec["prompt_mode"] is not None:
                prompt_candidates = [model_spec["prompt_mode"]]
            else:
                prompt_candidates = SAM_PROMPT_MODE_CANDIDATES

            for prompt_mode in prompt_candidates:
                candidate = (
                    model_root
                    / prompt_mode
                    / split_key
                    / "metrics_image_level.csv"
                )

                if candidate.exists():
                    return candidate, prompt_mode, folder_name

            fallback_candidates = sorted(
                model_root.glob(f"*/{split_key}/metrics_image_level.csv")
            )

            if fallback_candidates:
                path = fallback_candidates[0]
                prompt_mode = path.parents[1].name
                return path, prompt_mode, folder_name

        else:
            prompt_mode = model_spec["prompt_mode"]

            candidate = (
                model_root
                / prompt_mode
                / split_key
                / "metrics_image_level.csv"
            )

            if candidate.exists():
                return candidate, prompt_mode, folder_name

            fallback_candidates = sorted(
                model_root.glob(f"*/{split_key}/metrics_image_level.csv")
            )

            if fallback_candidates:
                path = fallback_candidates[0]
                prompt_mode = path.parents[1].name
                return path, prompt_mode, folder_name

    return None, None, None


def make_sample_id(df: pd.DataFrame):
    candidate_columns = [
        "image_name",
        "case_id",
        "mask_name",
        "merged_mask_name",
        "prediction_name",
    ]

    for column in candidate_columns:
        if column in df.columns:
            return df[column].astype(str).apply(lambda x: Path(x).stem)

    raise ValueError(
        f"Could not create sample_id. Available columns: {list(df.columns)}"
    )


def find_metric_column(df: pd.DataFrame, metric_name: str):
    lower_to_original = {
        str(column).lower(): column
        for column in df.columns
    }

    if metric_name.lower() in lower_to_original:
        return lower_to_original[metric_name.lower()]

    aliases = {
        "dice": [
            "dice_score",
            "dice_coefficient",
            "mean_dice",
            "f1",
            "f1_score",
        ],
        "iou": [
            "jaccard",
            "jaccard_index",
            "iou_score",
            "mean_iou",
        ],
        "recall": [
            "sensitivity",
            "true_positive_rate",
            "tpr",
            "recall_score",
        ],
        "precision": [
            "positive_predictive_value",
            "ppv",
            "precision_score",
        ],
    }

    for alias in aliases.get(metric_name.lower(), []):
        if alias in lower_to_original:
            return lower_to_original[alias]

    return None


def load_model_values(dataset_key: str, split_key: str, metric_name: str, model_spec: dict):
    metrics_path, prompt_mode, folder_name = find_metrics_file_for_model(
        dataset_key=dataset_key,
        split_key=split_key,
        model_spec=model_spec,
    )

    if metrics_path is None:
        print(
            f"WARNING: Missing file: "
            f"{dataset_key} | {split_key} | {metric_name} | {model_spec['display_name']}"
        )
        return None

    df = pd.read_csv(metrics_path)

    metric_column = find_metric_column(df, metric_name)

    if metric_column is None:
        print(
            f"WARNING: Missing metric column '{metric_name}': "
            f"{dataset_key} | {split_key} | {model_spec['display_name']} | {metrics_path}"
        )
        return None

    sample_id = make_sample_id(df)
    values = pd.to_numeric(df[metric_column], errors="coerce")

    out = pd.DataFrame(
        {
            "sample_id": sample_id,
            model_spec["display_name"]: values,
        }
    )

    out = out.dropna()

    return out


def load_group_wide_table(dataset_key: str, split_key: str, metric_name: str, group_name: str):
    model_tables = []

    for model_spec in MODEL_SPECS:
        if model_spec["group"] != group_name:
            continue

        model_df = load_model_values(
            dataset_key=dataset_key,
            split_key=split_key,
            metric_name=metric_name,
            model_spec=model_spec,
        )

        if model_df is None:
            continue

        model_tables.append(model_df.set_index("sample_id"))

    if len(model_tables) < 2:
        return None

    wide_df = pd.concat(model_tables, axis=1, join="inner")
    wide_df = wide_df.dropna(axis=0, how="any")
    wide_df = wide_df.reset_index()

    if len(wide_df) == 0:
        return None

    return wide_df


def holm_adjust(p_values):
    p_values = np.asarray(p_values, dtype=float)
    n = len(p_values)

    if n == 0:
        return np.array([])

    order = np.argsort(p_values)
    adjusted = np.empty(n, dtype=float)

    running_max = 0.0

    for rank, original_index in enumerate(order):
        multiplier = n - rank
        adjusted_value = p_values[original_index] * multiplier
        running_max = max(running_max, adjusted_value)
        adjusted[original_index] = min(running_max, 1.0)

    return adjusted


def wilcoxon_z_and_r_from_p(p_value: float, n: int):
    """
    Approximate Wilcoxon effect size r from two-sided p-value.

    r = |z| / sqrt(n)

    This is a common, simple reportable effect size for Wilcoxon tests.
    """

    if p_value is None or pd.isna(p_value):
        return np.nan, np.nan

    if n <= 0:
        return np.nan, np.nan

    if p_value <= 0:
        p_value = 1e-300

    if p_value >= 1:
        return 0.0, 0.0

    z_abs = abs(norm.ppf(p_value / 2.0))
    r = z_abs / np.sqrt(n)

    return z_abs, r


def summarize_and_test_group(
    wide_df: pd.DataFrame,
    dataset_key: str,
    split_key: str,
    metric_name: str,
    group_name: str,
):
    model_names = [
        column
        for column in wide_df.columns
        if column != "sample_id"
    ]

    summary_rows = []

    for model_name in model_names:
        values = pd.to_numeric(wide_df[model_name], errors="coerce").dropna()

        q1 = values.quantile(0.25)
        q3 = values.quantile(0.75)

        summary_rows.append(
            {
                "Dataset": dataset_key,
                "Split": split_key,
                "Metric": metric_name,
                "Group": group_name,
                "Model": model_name,
                "n": len(values),
                "Mean": values.mean(),
                "SD": values.std(),
                "Median": values.median(),
                "Q1": q1,
                "Q3": q3,
                "IQR": q3 - q1,
                "Mean ± SD": f"{values.mean():.4f} ± {values.std():.4f}",
                "Median [Q1, Q3]": f"{values.median():.4f} [{q1:.4f}, {q3:.4f}]",
            }
        )

    summary_df = pd.DataFrame(summary_rows)

    summary_df = summary_df.sort_values(
        by=["Mean", "Median"],
        ascending=False,
    ).reset_index(drop=True)

    summary_df.insert(
        summary_df.columns.get_loc("Model") + 1,
        "Rank",
        np.arange(1, len(summary_df) + 1),
    )

    best_model = summary_df.iloc[0]["Model"]

    test_rows = []
    raw_p_values = []
    raw_p_row_indices = []

    for index, row in summary_df.iterrows():
        model_name = row["Model"]

        if model_name == best_model:
            test_rows.append(
                {
                    "Best model": best_model,
                    "Compared with best": "Best model",
                    "Wilcoxon p": np.nan,
                    "Holm-adjusted p": np.nan,
                    "Significant after Holm": "",
                    "Mean difference vs best": np.nan,
                    "Median difference vs best": np.nan,
                    "Wilcoxon |z|": np.nan,
                    "Effect size r": np.nan,
                    "Interpretation": "Selected as best by highest mean",
                }
            )
            continue

        pair_df = wide_df[["sample_id", best_model, model_name]].copy()
        pair_df[best_model] = pd.to_numeric(pair_df[best_model], errors="coerce")
        pair_df[model_name] = pd.to_numeric(pair_df[model_name], errors="coerce")
        pair_df = pair_df.dropna()

        best_values = pair_df[best_model].to_numpy()
        other_values = pair_df[model_name].to_numpy()

        differences = best_values - other_values

        if len(pair_df) == 0:
            p_value = np.nan
            statistic = np.nan
            note = "No paired values"
        elif np.allclose(differences, 0):
            p_value = 1.0
            statistic = 0.0
            note = "No difference"
        else:
            try:
                statistic, p_value = wilcoxon(
                    best_values,
                    other_values,
                    alternative="two-sided",
                    zero_method="wilcox",
                )
                note = ""
            except Exception as error:
                statistic = np.nan
                p_value = np.nan
                note = str(error)

        z_abs, effect_r = wilcoxon_z_and_r_from_p(
            p_value=p_value,
            n=len(pair_df),
        )

        test_rows.append(
            {
                "Best model": best_model,
                "Compared with best": f"{best_model} vs {model_name}",
                "Wilcoxon p": p_value,
                "Holm-adjusted p": np.nan,
                "Significant after Holm": "",
                "Mean difference vs best": np.mean(differences),
                "Median difference vs best": np.median(differences),
                "Wilcoxon |z|": z_abs,
                "Effect size r": effect_r,
                "Interpretation": note,
            }
        )

        if not pd.isna(p_value):
            raw_p_values.append(p_value)
            raw_p_row_indices.append(index)

    adjusted_p_values = holm_adjust(raw_p_values)

    for adjusted_p, row_index in zip(adjusted_p_values, raw_p_row_indices):
        test_rows[row_index]["Holm-adjusted p"] = adjusted_p
        test_rows[row_index]["Significant after Holm"] = (
            "Yes" if adjusted_p < 0.05 else "No"
        )

    test_df = pd.DataFrame(test_rows)

    final_df = pd.concat(
        [
            summary_df.reset_index(drop=True),
            test_df.reset_index(drop=True),
        ],
        axis=1,
    )

    final_df = final_df[
        [
            "Dataset",
            "Split",
            "Metric",
            "Group",
            "Rank",
            "Model",
            "n",
            "Mean ± SD",
            "Median [Q1, Q3]",
            "Mean",
            "SD",
            "Median",
            "Q1",
            "Q3",
            "Best model",
            "Compared with best",
            "Mean difference vs best",
            "Median difference vs best",
            "Wilcoxon p",
            "Holm-adjusted p",
            "Significant after Holm",
            "Effect size r",
            "Wilcoxon |z|",
            "Interpretation",
        ]
    ]

    return final_df


def make_sheet_table(split_key: str, metric_name: str):
    all_rows = []

    for dataset_key in DATASETS:
        for group_name in GROUPS:
            print(
                f"Processing: split={split_key} | metric={metric_name} | "
                f"dataset={dataset_key} | group={group_name}"
            )

            wide_df = load_group_wide_table(
                dataset_key=dataset_key,
                split_key=split_key,
                metric_name=metric_name,
                group_name=group_name,
            )

            if wide_df is None:
                all_rows.append(
                    pd.DataFrame(
                        [
                            {
                                "Dataset": dataset_key,
                                "Split": split_key,
                                "Metric": metric_name,
                                "Group": group_name,
                                "Rank": "",
                                "Model": "Not enough aligned models found",
                                "n": "",
                                "Mean ± SD": "",
                                "Median [Q1, Q3]": "",
                                "Mean": "",
                                "SD": "",
                                "Median": "",
                                "Q1": "",
                                "Q3": "",
                                "Best model": "",
                                "Compared with best": "",
                                "Mean difference vs best": "",
                                "Median difference vs best": "",
                                "Wilcoxon p": "",
                                "Holm-adjusted p": "",
                                "Significant after Holm": "",
                                "Effect size r": "",
                                "Wilcoxon |z|": "",
                                "Interpretation": "Missing files, missing metric columns, or no shared sample IDs",
                            }
                        ]
                    )
                )
                continue

            result_df = summarize_and_test_group(
                wide_df=wide_df,
                dataset_key=dataset_key,
                split_key=split_key,
                metric_name=metric_name,
                group_name=group_name,
            )

            all_rows.append(result_df)

    if not all_rows:
        return pd.DataFrame()

    return pd.concat(all_rows, ignore_index=True)


def format_excel_workbook(writer):
    workbook = writer.book

    header_fill = "D9EAF7"
    sam_fill = "E8F6F8"
    trained_fill = "F3EAF8"
    best_fill = "FFF2CC"

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        ws.freeze_panes = "A2"

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor=header_fill)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        header_map = {
            cell.value: cell.column
            for cell in ws[1]
        }

        group_col = header_map.get("Group")
        rank_col = header_map.get("Rank")
        model_col = header_map.get("Model")
        p_col = header_map.get("Holm-adjusted p")
        effect_col = header_map.get("Effect size r")

        for row_idx in range(2, ws.max_row + 1):
            group_value = ws.cell(row=row_idx, column=group_col).value if group_col else ""

            if group_value == "SAM":
                row_fill = PatternFill("solid", fgColor=sam_fill)
            elif group_value == "Trained":
                row_fill = PatternFill("solid", fgColor=trained_fill)
            else:
                row_fill = None

            rank_value = ws.cell(row=row_idx, column=rank_col).value if rank_col else None

            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

                if row_fill is not None:
                    cell.fill = row_fill

                if rank_value == 1:
                    cell.fill = PatternFill("solid", fgColor=best_fill)
                    cell.font = Font(bold=True)

            if model_col:
                ws.cell(row=row_idx, column=model_col).alignment = Alignment(
                    horizontal="left",
                    vertical="center",
                )

            if p_col:
                ws.cell(row=row_idx, column=p_col).number_format = "0.0000"

            if effect_col:
                ws.cell(row=row_idx, column=effect_col).number_format = "0.0000"

        for col_idx in range(1, ws.max_column + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0

            for cell in ws[column_letter]:
                value = cell.value
                value_length = len(str(value)) if value is not None else 0
                max_length = max(max_length, value_length)

            ws.column_dimensions[column_letter].width = min(
                max(max_length + 2, 10),
                35,
            )

        ws.auto_filter.ref = ws.dimensions


def main():
    OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        for split_key in SPLITS:
            for metric_name in METRICS:
                sheet_name = f"{split_key}_{metric_name}"

                table_df = make_sheet_table(
                    split_key=split_key,
                    metric_name=metric_name,
                )

                table_df.to_excel(
                    writer,
                    sheet_name=sheet_name,
                    index=False,
                )

        format_excel_workbook(writer)

    print("\nSaved Excel workbook:")
    print(OUTPUT_XLSX)


if __name__ == "__main__":
    main()